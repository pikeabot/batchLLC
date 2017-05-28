import scrapy
import pandas as pd
from openpyxl import load_workbook
import re
from bs4 import BeautifulSoup
import logging

class BatchSpider(scrapy.Spider):
  name = "batch"
  base_url='http://search.sunbiz.org'

  def start_requests(self):
    xl = pd.ExcelFile('test_1.xlsx')
    sheet_names = xl.sheet_names
    df = xl.parse(sheet_names[0])

    #for owner_name in df['OWNER_NAME_1']:
    #for i in range(0, len(df['OWNER_NAME_1'])):
    for i in range(0, 10):
      self.logger.info('Business Name: {0}'.format(df['OWNER_NAME_1'][i]))
      owner_name = df['OWNER_NAME_1'][i]
      name_list = owner_name.split()
      if name_list[-1] in ['LLC', 'LTD', 'INC', 'LLLP', 'LP']:
          name_list=name_list[:-1]
      searchname=''.join(name_list)
      searchorder='%20'.join(name_list)
      url = 'http://search.sunbiz.org/Inquiry/CorporationSearch/SearchResults?inquiryType=EntityName&searchNameOrder={0}&searchTerm={1}'.format(searchname, searchorder)
      self.logger.info('Current Index: {0}'.format(i))
      yield scrapy.Request(url=url, callback=lambda response: self.parse_for_url(response, index=str(i)))


  def parse_for_url(self, response, index):
    url = response.xpath('//td[@class="large-width"]//@href').extract_first()
    url = self.base_url+url
    self.logger.info('URL Index: {0}'.format(index))
    yield scrapy.Request(url=url, callback=lambda response: self.parse_info(response, index=index))

  def parse_info(self, response, index):
    soup = None
    # Find out of an Authorized Person(s) sections exists
    detailedSections = response.xpath('//div[@class="detailSection"]')
    for section in detailedSections:
        if 'Authorized' in section.extract():
            soup = BeautifulSoup(section.extract(), "lxml")
            break

    if soup:
      persons = re.sub(r'Title.\S+', 'TitleAMBR', soup.body.get_text())
      persons = persons.split('TitleAMBR')
      persons.pop(0)  # Remove 'Authorized Person(s) Detail' title and other info before owners

      last_names =[]
      first_names = []
      addresses = []
      cities = []
      states = []
      zip_codes = []
      for person in persons:
        person = re.sub('[\r\n]+|\.|[\r\n]+', '', person)
        person = re.sub('  +', ';', person)
        info = person.split(';')
        last_names.append(info[0].split(',')[0])
        first_names.append(info[0].split(',')[1])
        addresses.append(info[1])
        city_state = re.findall(r'\S+, \S+', info[2])[0].split(',')
        cities.append(city_state[0])
        states.append(city_state[1])
        zip_codes.append(re.findall(r'\d+', info[2])[0])

        #names.append(i.lstrip().rstrip())
        wb = load_workbook(filename = 'test_1.xlsx')
        sheet_names = wb.get_sheet_names()
        ws = wb[sheet_names[0]]
        ws['F'+index] = ('\n').join(last_names)
        ws['G'+index] = ('\n').join(first_names)
        ws['R'+index] = ('\n').join(addresses)
        ws['S'+index] = ('\n').join(cities)
        ws['T'+index] = ('\n').join(states)
        ws['U'+index] = ('\n').join(zip_codes)
        wb.save('test_1.xlsx')

      self.logger.info('Index: {0}'.format(index))
      self.logger.info('person: {0}'.format(last_names))
    return



