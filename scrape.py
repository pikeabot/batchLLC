import sys
import urllib2
import argparse
import datetime
import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re
from bs4 import BeautifulSoup
import threading
import logging


class Scraper:
  def __init__(self, filename):
    self.LOG_FILENAME = 'logging.out'
    self.filename = filename
    self.row_count = self.get_row_count(filename)
    self.sheet0 = self.get_sheet0(filename)
    self.start_logging()

  def start_logging(self):
    logging.basicConfig(filename=self.LOG_FILENAME,level=logging.INFO)
    return

  def get_row_count(self, filename):
    # Open file for reading
    try:
      xl = pd.ExcelFile(filename)
      # Get the first sheet name
      sheet_names = xl.sheet_names
      df = xl.parse(sheet_names[0])
      return len(df['OWNER_NAME_1'])
    except:
      print('Error opening file')
      sys.exit(1)

  def get_sheet0(self, filename):
    # Open file for reading
    try:
      xl = pd.ExcelFile(filename)
      # Get the first sheet name
      return xl.sheet_names[0]
    except:
      print('Error opening file')
      sys.exit(1)

  '''
    This function uses beautiful soup to get the xml from the webpage url
  '''
  def soupify(self, url):
    user_agent = 'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_4; en-US) AppleWebKit/534.3 (KHTML, like Gecko) Chrome/6.0.472.63 Safari/534.3'
    headers = { 'User-Agent' : user_agent }
    req = urllib2.Request(url, None, headers)
    response = urllib2.urlopen(req)
    page = response.read()

    return BeautifulSoup(page, 'html.parser')

  '''
    Function that scrapes sunbiz.org for LLC owner information.
    It reads the business names off of an excel sheet and creates the query url to find the business url.
    Each business is on a separate thread to avoid race conditions. 
  '''
  def start_requests(self):

    threads = []

    # Open file for reading
    try:
      xl = pd.ExcelFile(self.filename)
      df = xl.parse(self.sheet0)
    except:
      print('Error opening file')
      sys.exit(1)

    # Loop through the list of LLC business names
    for i in range(0, self.row_count):
      try: 
        owner_name = df['OWNER_NAME_1'][i]
        print 'Checking for {}'.format(owner_name)
        name_list = owner_name.split()
        # Remove LLC, LTD, INC, LLLP  and LP from the name and create he business query url
        # This is done to conform with the query string format
        if name_list[-1] in ['LLC', 'LTD', 'INC', 'LLLP', 'LP']:
            name_list=name_list[:-1]
        searchname=''.join(name_list)
        searchorder='%20'.join(name_list)
        url = 'http://search.sunbiz.org/Inquiry/CorporationSearch/SearchResults?inquiryType=EntityName&searchNameOrder={0}&searchTerm={1}'.format(searchname, searchorder)
        # Create a thread to run the query and scraping for a particular business name
        # This is done to avoid race conditions for the http calls
        t = threading.Thread(target=self.get_llc_info, args=(url, self.filename, owner_name, i))
        threads.append(t)
        t.start()
        t.join()
      except Exception as e:
        logging.info('\n')
        logging.info(datetime.datetime.now())
        logging.exception(str(e))
  '''
    This function gets the results of querying for the business's name in start_requests(). The results
    are a list of links. The first link is assumed to be the best match and is verified to be close to the 
    original business name. If a good match, the link is followed and then that business's sunbiz.org site is 
    scraped for the owner info. If an owner is found to exist, then the info is parsed and saved to an excel file.
  '''
  def get_llc_info(self, url, file_name, owner_name, i):
    # Get the results of the business name query and find the first link
    soup = self.soupify(url)
    search_result = soup.find('td', class_='large-width')
    # Use fuzzy string matching to determine how similar the business name is to the name in the link
    owner_str = re.sub('[^A-Za-z0-9]+', '', owner_name)
    search_str = re.sub('[^A-Za-z0-9]+', '', search_result.get_text())
    if fuzz.partial_ratio(owner_str, search_str) > 75:
      try:
        # Good match then go to business site
        llc_soup =  self.soupify('http://search.sunbiz.org' + search_result.a.get('href'))
      except Exception as e: 
        logging.info('\n')
        logging.info(str(datetime.datetime.now()) +':   Error reading url for {0} [{1}]'.format(owner_name, str(i+1)))
        logging.exception(str(e)) 
      # Find the section with Authorized Person(s) Information.
      # There may be more than one owner
      detailed_sections = llc_soup.find_all("div", "detailSection")
      for section in detailed_sections:
        if 'Authorized' in section.get_text():
            llc_info = section.get_text()
            persons = re.sub(r'Title.\S+.*\S*', 'TitleAMBR', llc_info)
            persons = persons.split('TitleAMBR')
            persons.pop(0)  # Remove 'Authorized Person(s) Detail' title and other info before owners

            last_names =[]
            first_names = []
            addresses = []
            cities = []
            states = []
            zip_codes = []
            # Parse each owner
            # Info is assumed to be:
            #   last name, first name
            #   street address
            #   (optional) apt/suite #
            #   city, state zip code
            for person in persons:  
              person = re.sub('[\r\n]+|\.|[\r\n]+', '', person)
              person = re.sub('  +', ';', person)
              info = person.split(';')

              try:
                if ',' in info[0]:
                  last_names.append(info[0].split(',')[0])
                  first_names.append(info[0].split(',')[1])
                else:
                  last_names.append(info[0])
                  first_names.append('\x20')
              except Exception as e: 
                logging.info('\n')
                logging.info(str(datetime.datetime.now()) +':   Error parsing first and last names for {0} [{1}]'.format(owner_name, str(i+1)))
                logging.exception(str(e))

              try:
                # Checks if there is an apt/suite # line
                if len(info) > 3:
                  addresses.append(' '.join([info[1], info[2]]))
                else:
                  addresses.append(info[1])
              except Exception as e:
                logging.info('\n')
                logging.info(str(datetime.datetime.now()) +': Error parsing street address for {0} [{1}]'.format(owner_name, str(i+1)))
                logging.exception(str(e))
              try: 
                last_info_line = info[-1]
                city_state = re.findall(r'\S+, \S+', last_info_line)[0].split(',')
                cities.append(city_state[0])
                states.append(city_state[1])
              except Exception as e: 
                logging.info('\n')
                logging.info(str(datetime.datetime.now()) +': Error parsing for city and state for {0} [{1}]'.format(owner_name, str(i+1)))
                logging.exception(str(e))
                cities.append(last_info_line)
                states.append('\x20')
              try:
                zip_codes.append(re.findall(r'\d+', last_info_line)[0])
              except Exception as e: 
                logging.info('\n')
                logging.info(str(datetime.datetime.now()) +': Error parsing for zip code for {0} [{1}]'.format(owner_name, str(i+1)))
                logging.exception(str(e))
                zip_codes.append('\x20')
            try:
              # Write info to excel file
              wb = load_workbook(filename = file_name)
              sheet_names = wb.get_sheet_names()
              ws = wb[sheet_names[0]]
              index = str(i+2)
              ws['F'+index] = ('\n').join(last_names)
              ws['F'+index].alignment = Alignment(wrapText=True)
              ws['G'+index] = ('\n').join(first_names)
              ws['G'+index].alignment = Alignment(wrapText=True)
              ws['R'+index] = ('\n').join(addresses)
              ws['R'+index].alignment = Alignment(wrapText=True)
              ws['S'+index] = ('\n').join(cities)
              ws['S'+index].alignment = Alignment(wrapText=True)
              ws['T'+index] = ('\n').join(states)
              ws['T'+index].alignment = Alignment(wrapText=True)
              ws['U'+index] = ('\n').join(zip_codes)
              ws['U'+index].alignment = Alignment(wrapText=True)
              wb.save(file_name)
            except Exception as e:
                logging.info('\n')
                logging.info(str(datetime.datetime.now()) +': Error writing data to excel file for {0} [{1}]'.format(owner_name, str(i+1)))
                logging.exception(str(e))  
            break  
    
    return

  def pretty(self):
    
    try:
      # Write info to excel file
      wb = load_workbook(filename = self.filename)
      filename_updated = self.sheet0 + ' (updated)'
      if not filename_updated in wb.sheetnames:
        ws_pretty = wb.create_sheet(filename_updated, 1)
        wb.save(self.filename)

      # Start reading lines and adding rows to the pandas dataframe for the case
      # where there are multiple owners
      # read dataframe from sheet
      df = pd.read_excel(self.filename, sheet_name=self.sheet0)
      # create a temporary dataframe to store the updated info
      df_pretty = pd.DataFrame()

      for row in df.iterrows():
        # Check for multiple owners
        if not isinstance(row[1]['Owner_Last Name'], float) and '\n' in row[1]['Owner_Last Name']:
          last_names = row[1]['Owner_Last Name'].split('\n')
          first_names = row[1]['Owner_First Name'].split('\n')
          addresses = row[1]['OWNER_ADDRESS'].split('\n')
          cities = row[1]['OWNER_CITY'].split('\n')
          states = row[1]['OWNER_STATE'].split('\n')
          zip_codes = row[1]['OWNER_ZIPCODE'].split('\n')
          # start adding lines
          for i in range(0, len(row[1]['Owner_Last Name'].split('\n'))):
            df_temp = row[1]
            df_temp['Owner_Last Name'] = last_names[i]
            if not isinstance(row[1]['Owner_First Name'], float) or row[1]['Owner_First Name'] != '': 
              df_temp['Owner_First Name'] = first_names[i] 
            df_temp['OWNER_ADDRESS'] = addresses[i]
            df_temp['OWNER_CITY'] = cities[i]
            df_temp['OWNER_STATE'] = states[i]
            df_temp['OWNER_ZIPCODE'] = zip_codes[i] 
            df_pretty = pd.concat([df_pretty, df_temp.to_frame().T]) 
        # no multiple owners -copy original line
        else:
          df_pretty = pd.concat([df_pretty, row[1].to_frame().T])  
      # write to file
      writer = pd.ExcelWriter(self.filename,  engine='openpyxl')
      writer.book = wb
      writer.sheets = dict((ws.title,ws) for ws in wb.worksheets)
      df_pretty.to_excel(writer, sheet_name=filename_updated, index=False)    
      writer.save()    
    except Exception as e:
        logging.info('\n')
        logging.info('error at {0}'.format(last_names[i]))
        logging.info(row[1]['Owner_First Name'])
        logging.exception(str(e))  

    # Format sheet -hide necessary columns and widen other columns
    try:
      # Write info to excel file
      wb = load_workbook(filename = self.filename)
      ws = wb[filename_updated]
      for col in ['A', 'B', 'C', 'D', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'x']:
        ws.column_dimensions[col].hidden= True
      for col in ['E', 'V']:
        ws.column_dimensions[col].width = 43
      for col in ['F', 'G', 'R', 'S', 'T', 'U', 'W']:
        ws.column_dimensions[col].width = 13
      wb.remove_sheet(wb[self.sheet0])
      ws.title = self.sheet0
      wb.save(self.filename)
    except Exception as e:
        logging.info('\n')
        logging.info(str(datetime.datetime.now()) +': Error updating worksheet format')
        logging.exception(str(e))  

if __name__ == '__main__':
  # Get the excel file from the command line
  parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
  parser.add_argument('--file', help='name (and path if necessary) to the excel file', required=True)
  args =  parser.parse_args()

  scraper=Scraper(args.file)
  scraper.start_requests()
  scraper.pretty()
  print 'Script is finished!'